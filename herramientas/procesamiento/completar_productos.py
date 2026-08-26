"""
completar_productos.py
------------------------
Completa productos.json con los productos del inventario que aún NO están
en el catálogo (376 de 805 SKUs faltan actualmente). NO toca ni sobrescribe
los 429 productos que ya tienes -- solo agrega los que faltan.

Cada producto nuevo se crea con imagen GENÉRICA por categoría (segura, nunca
imagen incorrecta) usando la misma convención que ya usa tu script.js.

USO:
    python completar_productos.py

Genera:
    productos.json                 -> actualizado (429 existentes + 376 nuevos = 805)
    productos_agregados_reporte.csv -> lista de los nuevos SKUs agregados, para que
                                        sepas cuáles priorizar para foto real
"""

import json
import re
import csv
import openpyxl

EXCEL_PATH = "INVENTARIO.xlsx"  # ajusta el nombre exacto de tu archivo si difiere
JSON_PATH = "productos.json"
HOJA = "INVENTARIO"

# Mismas reglas de script.js (IMAGENES_MEDICAMENTOS_REGLAS), para asignar
# imagen específica si el nombre coincide con un medicamento ya curado.
REGLAS_IMAGEN = [
    (r"enalapril", "imagenesmedimentos/enalapril20mg.jpg"),
    (r"losartan", "imagenesmedimentos/losartan50mg.png"),
    (r"acetaminofen", "imagenesmedimentos/Acetaminofen500mg.png"),
    (r"aspirina|cardioaspirina", "imagenesmedimentos/aspirina500mg.png"),
    (r"dolex", "imagenesmedimentos/dolex.jpg"),
    (r"hioscina|dipirona|buscapina", "imagenesmedimentos/buscapnacompositum.jpg"),
    (r"amoxicilina|amoxidal", "imagenesmedimentos/amoxicilina500mg.jpg"),
    (r"cefalexina", "imagenesmedimentos/cefalexina500mg.png"),
    (r"advil", "imagenesmedimentos/advil.jpg"),
    (r"noxpirin", "imagenesmedimentos/nospirina.jpg"),
    (r"loratadina", "imagenesmedimentos/loratadina10mg.png"),
    (r"diclofenaco|dormex", "imagenesmedimentos/diclofenacogel.png"),
    (r"ibuprofeno", "imagenesmedimentos/Ibuprofeno800mg.png"),
    (r"naproxeno", "imagenesmedimentos/naproxeno500mg.png"),
    (r"similac", "imagenesmedimentos/similac1.jpg"),
    (r"enfamil", "imagenesmedimentos/enfamilpremium.jpg"),
    (r"pa[ñn]al|huggies", "imagenesmedimentos/pañaleshuggies.jpg"),
    (r"winny", "imagenesmedimentos/pañaleswinny.jpg"),
    (r"electrolit", "imagenesmedimentos/electrolit.png"),
    (r"pedialyte", "imagenesmedimentos/pedialyte.png"),
    (r"omeprazol", "imagenesmedimentos/omeprazol.png"),
    (r"diosmectita", "imagenesmedimentos/diosmectita.png"),
    (r"salbutamol", "imagenesmedimentos/salbutamol100mcg.png"),
    (r"colgate", "imagenesmedimentos/colgatetripleaccion.jpg"),
    (r"hidrocortisona", "imagenesmedimentos/hidrocortisonacrema.png"),
    (r"ketoconazol", "imagenesmedimentos/ketoconazolcrema.jpg"),
    (r"glibenclamida", "imagenesmedimentos/glibenclamida5mg.jpg"),
    (r"metformina", "imagenesmedimentos/metformina850mg.png"),
    (r"desodorante|gillette", "imagenesmedimentos/desoderantegillette82gx2.png"),
    (r"centrum", "imagenesmedimentos/centrum.jpg"),
    (r"scott", "imagenesmedimentos/scott.jpg"),
    (r"vita\s*c|vitac", "imagenesmedimentos/vitac500mg.jpg"),
]

# Palabras clave -> categoría (para los 376 nuevos que no tienen categoría asignada)
CATEGORIA_KEYWORDS = [
    (r"amoxicilina|cefalexina|amoxidal|azitromicina|claritromicina|ciprofloxacino", "Antibióticos"),
    (r"ibuprofeno|naproxeno|diclofenaco|dormex", "Antiinflamatorios"),
    (r"acetaminofen|aspirina|cardioaspirina|dolex|dipirona|buscapina|advil|nospirina|noxpirin", "Analgésicos"),
    (r"loratadina|antigripal|gripa|descongestionante", "Antigripales"),
    (r"vitamina|centrum|vitac|complejo b|multivitaminico", "Vitaminas"),
    (r"pañal|huggies|winny|similac|enfamil|biberon|leche.*infantil", "Bebés"),
    (r"gaseosa|jugo|agua|electrolit|pedialyte|bebida|gatorade", "Bebidas"),
    (r"crema|shampoo|jabon|desodorante|gillette|colgate|cepillo|protector solar|talco", "Cuidado personal"),
    (r"losartan|enalapril|metformina|glibenclamida", "Adulto mayor"),
]

CATEGORIA_IMAGENES_GENERICAS = {
    'Antibióticos': 'imagenesmedimentos/genericas/antibioticos.svg',
    'Antigripales': 'imagenesmedimentos/genericas/antigripales.svg',
    'Vitaminas': 'imagenesmedimentos/genericas/vitaminas.svg',
    'Antiinflamatorios': 'imagenesmedimentos/genericas/antiinflamatorios.svg',
    'Bebidas': 'imagenesmedimentos/genericas/bebidas.svg',
    'Cuidado personal': 'imagenesmedimentos/genericas/cuidado-personal.svg',
    'Bebés': 'imagenesmedimentos/genericas/bebes.svg',
    'Adulto mayor': 'imagenesmedimentos/genericas/adulto-mayor.svg',
    'Analgésicos': 'imagenesmedimentos/genericas/otros.svg',  # no existe analgesicos.svg en tu árbol; ajusta si la tienes
    'Otros': 'imagenesmedimentos/genericas/otros.svg',
}


def clasificar_categoria(texto):
    t = texto.lower()
    for patron, categoria in CATEGORIA_KEYWORDS:
        if re.search(patron, t):
            return categoria
    return "Otros"


def imagen_por_nombre(texto):
    t = texto.lower()
    for patron, ruta in REGLAS_IMAGEN:
        if re.search(patron, t):
            return ruta
    return None


def main():
    with open(JSON_PATH, "r", encoding="utf-8-sig") as f:
        productos = json.load(f)

    ids_existentes = {p["id"] for p in productos}

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[HOJA]
    encabezados = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {n: i for i, n in enumerate(encabezados)}

    nuevos = []
    vistos_cod = set()

    for fila in ws.iter_rows(min_row=2, values_only=True):
        cod = fila[idx["COD"]]
        descripcion = (fila[idx["DESCRIPCION"]] or "").strip()
        if not cod or not descripcion:
            continue
        cod = str(cod).strip()
        if cod in vistos_cod:
            continue
        vistos_cod.add(cod)

        id_prod = f"prod-{cod}"
        if id_prod in ids_existentes:
            continue  # ya existe, no lo tocamos

        precio = fila[idx.get("PVP_F", idx.get("PRE"))] or 0
        categoria = clasificar_categoria(descripcion)
        imagen = imagen_por_nombre(descripcion) or CATEGORIA_IMAGENES_GENERICAS.get(categoria, CATEGORIA_IMAGENES_GENERICAS["Otros"])

        nuevo = {
            "id": id_prod,
            "nombre": descripcion[:30],
            "precio": int(precio) if precio else 0,
            "categoria": categoria,
            "descripcion": descripcion,
            "imagen": imagen,
            "stock": 1,
            "disponible": True,
        }
        productos.append(nuevo)
        nuevos.append(nuevo)

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(productos, f, ensure_ascii=False, indent=2)

    with open("productos_agregados_reporte.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["id", "nombre", "categoria", "imagen"])
        w.writeheader()
        for n in nuevos:
            w.writerow({"id": n["id"], "nombre": n["nombre"], "categoria": n["categoria"], "imagen": n["imagen"]})

    print(f"Productos existentes conservados: {len(ids_existentes)}")
    print(f"Productos nuevos agregados:       {len(nuevos)}")
    print(f"Total final en productos.json:    {len(productos)}")
    print("-> productos.json actualizado")
    print("-> productos_agregados_reporte.csv (revisa la categoría/imagen asignada a los nuevos)")


if __name__ == "__main__":
    main()
